import base64
import io
import json
import threading
import traceback
from datetime import date, datetime

import daff
import openpyxl
from flask import Blueprint, session, current_app, g, render_template, request
from flask_github import GitHub
from openpyxl.styles import Font, PatternFill

from ..OntologyDataStore import OntologyDataStore
from ..SpreadsheetSearcher import SpreadsheetSearcher
from ..guards.verify_login import verify_logged_in
from ..services.ConfigurationService import ConfigurationService
from ..utils.github import get_spreadsheet, get_csv, create_branch

bp = Blueprint("edit", __name__, template_folder="../templates")


@bp.route('/edit/<repo_key>/<path:folder>/<spreadsheet>')
@verify_logged_in
def edit(repo_key, folder, spreadsheet, github: GitHub, ontodb: OntologyDataStore, config: ConfigurationService):
    if session.get('label') is None:
        go_to_row = ""
    else:
        go_to_row = session.get('label')
        session.pop('label', None)

    if session.get('type') is None:
        type = ""
    else:
        type = session.get('type')
        session.pop('type', None)
    repository = config.get(repo_key)
    (file_sha, rows, header) = get_spreadsheet(github, repository.full_name, folder, spreadsheet)
    if g.user.github_login in config.app_config['USERS']:
        user_initials = config.app_config['USERS'][g.user.github_login]["initials"]
    else:
        current_app.logger.info(f"The user {g.user.github_login} has no known metadata")
        user_initials = g.user.github_login[0:2]
    # Build suggestions data:
    if repo_key not in ontodb.releases or date.today() > ontodb.releasedates[repo_key]:
        ontodb.parseRelease(repo_key)
    suggestions = ontodb.getReleaseLabels(repo_key)
    suggestions = list(dict.fromkeys(suggestions))

    breadcrumb_segments = [repo_key, *folder.split("/"), spreadsheet]

    return render_template('edit.html',
                           login=g.user.github_login,
                           user_initials=user_initials,
                           all_initials=[v["initials"] for v in config.app_config["USERS"].values()],
                           repo_name=repo_key,
                           folder=folder,
                           spreadsheet_name=spreadsheet,
                           breadcrumb=[{"name": s, "path": "repo/" + "/".join(breadcrumb_segments[:i + 1])} for i, s in
                                       enumerate(breadcrumb_segments)],

                           header=json.dumps(header),
                           rows=json.dumps(rows),
                           file_sha=file_sha,
                           go_to_row=go_to_row,
                           type=type,
                           suggestions=json.dumps(suggestions)
                           )


@bp.route('/download_spreadsheet', methods=['POST'])
@verify_logged_in
def download_spreadsheet(github: GitHub, config: ConfigurationService):
    repo_key = request.form.get("repo_key")
    folder = request.form.get("folder")
    spreadsheet = request.form.get("spreadsheet")
    repository = config.get(repo_key)
    url = github.get(f"repos/{repository.full_name}/contents/{folder}/{spreadsheet}")
    download_url = url['download_url']
    current_app.logger.debug(f"Downloading spreadsheet from {download_url}")
    return (json.dumps({"message": "Success",
                        "download_url": download_url}), 200)


@bp.route('/save', methods=['POST'])
@verify_logged_in
def save(searcher: SpreadsheetSearcher, github: GitHub, config: ConfigurationService):
    repo_key = request.form.get("repo_key")
    folder = request.form.get("folder")
    spreadsheet = request.form.get("spreadsheet")
    row_data = request.form.get("rowData")
    initial_data = request.form.get("initialData")
    file_sha = request.form.get("file_sha").strip()
    commit_msg = request.form.get("commit_msg")
    commit_msg_extra = request.form.get("commit_msg_extra")
    overwrite = False
    overwriteVal = request.form.get("overwrite")
    if overwriteVal == "true":
        overwrite = True

    id_suffix = ""

    repository = config.get(repo_key)
    default_branch = repository.main_branch
    restart = False  # for refreshing the sheet (new ID's)
    try:
        initial_data_parsed = json.loads(initial_data)
        row_data_parsed = json.loads(row_data)
        # Sort based on label
        # What if 'Label' column not present?
        initial_data_parsed = sorted(initial_data_parsed, key=lambda k: k.get('Label', None) or "")

        first_row = row_data_parsed[0]
        header = [k for k in first_row.keys()]
        del header[0]
        # Sort based on label
        # What if 'Label' column not present?
        if 'Label' in first_row:
            row_data_parsed = sorted(row_data_parsed, key=lambda k: k['Label'] if k['Label'] else "")
        elif 'Relationship' in first_row:
            row_data_parsed = sorted(row_data_parsed, key=lambda k: k['Relationship'] if k['Relationship'] else "")

            id_suffix = "R"  # For relations add a 'R' as a suffix
        else:
            current_app.logger.warning(
                "While saving: No Label column present, so not sorting this.")  # do we need to sort - yes, for diff!

        current_app.logger.debug(f"Got file_sha: {file_sha}")

        next_id = searcher.get_next_id(repo_key + id_suffix)

        wb = openpyxl.Workbook()
        sheet = wb.active

        for c in range(len(header)):
            sheet.cell(row=1, column=c + 1).value = header[c]
            sheet.cell(row=1, column=c + 1).font = Font(size=12, bold=True)
        for r in range(len(row_data_parsed)):
            row = [v for v in row_data_parsed[r].values()]
            del row[0]  # Tabulator-added ID column
            for c in range(len(header)):
                sheet.cell(row=r + 2, column=c + 1).value = row[c]
                # Set row background colours according to 'Curation status'
                # These should be kept in sync with those used in edit screen
                # TODO add to config
                # What if "Curation status" not present?
                if 'Curation status' in first_row:
                    if row[header.index("Curation status")] == "Discussed":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="ffe4b5", fill_type="solid")
                    elif row[header.index("Curation status")] == "Ready":  # this is depreciated
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="98fb98", fill_type="solid")
                    elif row[header.index("Curation status")] == "Proposed":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="ffffff", fill_type="solid")
                    elif row[header.index("Curation status")] == "Pre-proposed":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="ebfad0", fill_type="solid")
                    elif row[header.index("Curation status")] == "To Be Discussed":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="eee8aa", fill_type="solid")
                    elif row[header.index("Curation status")] == "In Discussion":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="fffacd", fill_type="solid")
                    elif row[header.index("Curation status")] == "Published":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="7fffd4", fill_type="solid")
                    elif row[header.index("Curation status")] == "Obsolete":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="2f4f4f", fill_type="solid")

            # Generate identifiers:
            if 'ID' in first_row:
                if not row[header.index("ID")]:  # blank
                    label, parent, definition = None, None, None
                    if {'Label', 'Parent', 'Definition'} <= set(first_row):  # make sure we have the right sheet
                        label = row[header.index("Label")]
                        parent = row[header.index("Parent")]
                        definition = row[header.index("Definition")]
                    elif {'Relationship', 'Parent relationship', 'Definition'} <= set(first_row):
                        label = row[header.index("Relationship")]
                        parent = row[header.index("Parent relationship")]
                        definition = row[header.index("Definition")]

                    if label and parent and definition:
                        # generate ID here:
                        nextIdStr = str(next_id)
                        next_id += 1
                        fill_num = repository.id_digits
                        if repo_key == "BCIO":
                            fill_num = fill_num - 1
                        else:
                            fill_num = fill_num
                        id = repo_key.upper() + id_suffix + ":" + nextIdStr.zfill(fill_num)
                        new_id = id
                        row_data_parsed[r]["ID"] = new_id
                        for c in range(len(header)):
                            if c == 0:
                                restart = True
                                sheet.cell(row=r + 2, column=c + 1).value = new_id

        # Create version for saving
        spreadsheet_stream = io.BytesIO()
        wb.save(spreadsheet_stream)

        # base64_bytes = base64.b64encode(sample_string_bytes)
        base64_bytes = base64.b64encode(spreadsheet_stream.getvalue())
        base64_string = base64_bytes.decode("ascii")

        # Create a new branch to commit the change to (in case of simultaneous updates)
        branch = f"{g.user.github_login}_{datetime.utcnow().strftime('%Y-%m-%d_%H%M%S')}"
        create_branch(github, repository.full_name, branch, default_branch)

        current_app.logger.debug("About to get latest version of the spreadsheet file %s",
                                 f"repos/{repository.full_name}/contents/{folder}/{spreadsheet}")
        # Get the sha for the file
        (new_file_sha, new_rows, new_header) = get_spreadsheet(github, repository.full_name, folder, spreadsheet)

        # Commit changes to branch (replace code with sheet)
        data = {"message": commit_msg, "content": base64_string, "branch": branch, "sha": new_file_sha}
        current_app.logger.debug("About to commit file to branch %s",
                                 f"repos/{repository.full_name}/contents/{folder}/{spreadsheet}")
        response = github.put(f"repos/{repository.full_name}/contents/{folder}/{spreadsheet}", data=data)
        if not response:
            raise Exception(
                f"Unable to commit addition of {spreadsheet} to branch {branch} in {repository.full_name}"
            )

        # Create a PR for the change
        current_app.logger.debug("About to create PR from branch", )
        response = github.post(
            f"repos/{repository.full_name}/pulls",
            data={
                "title": commit_msg,
                "head": branch,
                "base": default_branch,
                "body": commit_msg_extra
            },
        )
        if not response:
            raise Exception(f"Unable to create PR for branch {branch} in {repository.full_name}")
        pr_info = response['html_url']

        # Do not merge automatically if this file was stale as that will overwrite the other changes

        if new_file_sha != file_sha and not overwrite:
            current_app.logger.info("PR created and must be merged manually as repo file had changed")

            # Get the changes between the new file and this one:
            merge_diff, merged_table = getDiff(row_data_parsed, new_rows, new_header, initial_data_parsed)
            # update rows for comparison:
            (file_sha3, rows3, header3) = get_spreadsheet(github, repository.full_name, folder, spreadsheet)
            # todo: delete transient branch here? Github delete code is a test for now.
            # Delete the branch again
            current_app.logger.debug("About to delete branch", f"repos/{repository.full_name}/git/refs/heads/{branch}")
            response = github.delete(
                f"repos/{repository.full_name}/git/refs/heads/{branch}")
            if not response:
                raise Exception(f"Unable to delete branch {branch} in {repository.full_name}")
            return (
                json.dumps({
                    'Error': 'Your change was submitted to the repository but could not be automatically merged due '
                             'to a conflict. You can view the change <a href="'
                             + pr_info + '" target = "_blank" >here </a>. ', "file_sha_1": file_sha,
                    "file_sha_2": new_file_sha, "pr_branch": branch, "merge_diff": merge_diff,
                    "merged_table": json.dumps(merged_table),
                    "rows3": rows3, "header3": header3}), 300  # 400 for missing REPO
            )
        else:
            # Merge the created PR
            current_app.logger.debug("About to merge created PR")
            response = github.post(
                f"repos/{repository.full_name}/merges",
                data={
                    "head": branch,
                    "base": default_branch,
                    "commit_message": commit_msg
                },
            )
            if not response:
                raise Exception(f"Unable to merge PR from branch {branch} in {repository.full_name}")

            # Delete the branch again
            current_app.logger.debug("About to delete branch %s",
                                     f"repos/{repository.full_name}/git/refs/heads/{branch}")
            response = github.delete(
                f"repos/{repository.full_name}/git/refs/heads/{branch}")
            if not response:
                raise Exception(f"Unable to delete branch {branch} in {repository.full_name}")

        current_app.logger.info("Save succeeded.")
        # Update the search index for this file ASYNCHRONOUSLY (don't wait)
        thread = threading.Thread(target=searcher.update_index,
                                  args=(repo_key, folder, spreadsheet, header, row_data_parsed))
        thread.daemon = True  # Daemonize thread
        thread.start()  # Start the execution

        # Get the sha AGAIN for the file
        response = github.get(f"repos/{repository.full_name}/contents/{folder}/{spreadsheet}")
        if not response or "sha" not in response:
            raise Exception(
                f"Unable to get the newly updated SHA value for {spreadsheet} in {repository.full_name}/{folder}"
            )
        new_file_sha = response['sha']
        if restart:  # todo: does this need to be anywhere else also?
            return (json.dumps({"message": "Success",
                                "file_sha": new_file_sha}), 360)
        else:
            return (json.dumps({"message": "Success",
                                "file_sha": new_file_sha}), 200)

    except Exception as err:
        current_app.logger.error(err)
        traceback.print_exc()
        return (
            json.dumps({"message": "Failed",
                        "Error": format(err)}),
            400,
        )


# todo: use this function to compare initial spreadsheet to server version - check for updates?
@bp.route("/checkForUpdates", methods=["POST"])
def checkForUpdates(github: GitHub, config: ConfigurationService):
    if request.method == "POST":
        repo_key = request.form.get("repo_key")
        folder = request.form.get("folder")
        spreadsheet = request.form.get("spreadsheet")
        old_sha = request.form.get("file_sha")
        repository = config.get(repo_key)
        spreadsheet_file = github.get(
            f'repos/{repository.full_name}/contents/{folder}/{spreadsheet}'
        )
        file_sha = spreadsheet_file['sha']
        if old_sha == file_sha:
            return json.dumps({"message": "Success"}), 200
        else:
            return json.dumps({"message": "Fail"}), 200


@bp.route("/generate", methods=["POST"])
@verify_logged_in
def generate(searcher: SpreadsheetSearcher, config: ConfigurationService):
    if request.method == "POST":
        repo_key = request.form.get("repo_key")
        repository = config.get(repo_key)
        rowData = json.loads(request.form.get("rowData"))
        values = {}
        ids = {}
        for row in rowData:
            nextIdStr = str(searcher.get_next_id(repo_key))
            fill_num = repository.id_digits
            if repo_key == "BCIO":
                fill_num = fill_num - 1
            else:
                fill_num = fill_num
            id = repo_key.upper() + ":" + nextIdStr.zfill(fill_num)
            ids["ID" + str(row['id'])] = str(row['id'])
            values["ID" + str(row['id'])] = id
        return json.dumps({"message": "idlist", "IDs": ids, "values": values})  # need to return an array
    return 'success'


@bp.route("/validate", methods=["POST"])
@verify_logged_in
def verify():
    if request.method == "POST":
        cell = json.loads(request.form.get("cell"))
        column = json.loads(request.form.get("column"))
        rowData = json.loads(request.form.get("rowData"))
        headers = json.loads(request.form.get("headers"))
        table = json.loads(request.form.get("table"))
        # check for blank cells under conditions first:
    blank = {}
    unique = {}
    returnData, uniqueData = checkBlankMulti(1, blank, unique, cell, column, headers, rowData, table)
    if len(returnData) > 0 or len(uniqueData) > 0:
        return json.dumps({"message": "fail", "values": returnData, "unique": uniqueData})
    return 'success'


def checkBlankMulti(current, blank, unique, cell, column, headers, rowData, table):
    for index, (key, value) in enumerate(
            rowData.items()):  # todo: really, we need to loop here, surely there is a faster way?
        if index == current:
            if key in ["Label", "Definition", "Parent", "Sub-ontology", "Curation status"]:
                if key == "Definition" or key == "Parent":
                    status = rowData.get("Curation status")  # check for "Curation status"
                    if status:
                        if rowData["Curation status"] == "Proposed" or rowData["Curation status"] == "External":
                            pass
                        else:
                            if value.strip() == "":
                                blank.update({key: value})
                    else:
                        pass  # no "Curation status" column
                else:
                    if value.strip() == "":
                        blank.update({key: value})
                    else:
                        pass
            if key == "Label" or key == "ID" or key == "Definition":
                if checkNotUnique(value, key, headers, table):
                    unique.update({key: value})
    # go again:
    current = current + 1
    if current >= len(rowData):
        return blank, unique
    return checkBlankMulti(current, blank, unique, cell, column, headers, rowData, table)


def checkNotUnique(cell, column, headers, table):
    counter = 0
    cellStr = cell.strip()
    if cellStr == "":
        return False
    # if Label, ID or Definition column, check cell against all other cells in the same column and return true if same
    for r in range(len(table)):
        row = [v for v in table[r].values()]
        del row[0]  # remove extra numbered "id" column
        for c in range(len(headers)):
            if headers[c] == "ID" and column == "ID":
                if row[c].strip() == cellStr:
                    counter += 1
                    if counter > 1:  # more than one of the same
                        return True
            if headers[c] == "Label" and column == "Label":
                if row[c].strip() == cellStr:
                    counter += 1
                    if counter > 1:
                        return True
            if headers[c] == "Definition" and column == "Definition":
                if row[c].strip() == cellStr:
                    counter += 1
                    if counter > 1:
                        return True
    return False


def getDiff(row_data_1, row_data_2, row_header, row_data_3):  # (1saving, 2server, header, 3initial)

    # print(f'the type of row_data_3 is: ')
    # print(type(row_data_3))

    # sort out row_data_1 format to be the same as row_data_2
    new_row_data_1 = []
    for k in row_data_1:
        dictT = {}
        for key, val, item in zip(k, k.values(), k.items()):
            if key != "id":
                if val == "":
                    val = None
                # add to dictionary:
                dictT[key] = val
        # add to list:
        new_row_data_1.append(dictT)

        # sort out row_data_3 format to be the same as row_data_2
    new_row_data_3 = []
    for h in row_data_3:
        dictT3 = {}
        for key, val, item in zip(h, h.values(), h.items()):
            if key != "id":
                if val == "":
                    val = None
                # add to dictionary:
                dictT3[key] = val
        # add to list:
        new_row_data_3.append(dictT3)

    row_data_combo_1 = [row_header]
    row_data_combo_2 = [row_header]
    row_data_combo_3 = [row_header]

    row_data_combo_1.extend(
        [list(r.values()) for r in new_row_data_1])  # row_data_1 has extra "id" column for some reason???!!!
    row_data_combo_2.extend([list(s.values()) for s in row_data_2])
    row_data_combo_3.extend([list(t.values()) for t in new_row_data_3])

    # checking:
    # print(f'row_header: ')
    # print(row_header)
    # print(f'row_data_1: ')
    # print(row_data_1)
    # print(f'row_data_2: ')
    # print(row_data_2)
    # print(f'combined 1: ')
    # print(row_data_combo_1)
    # print(f'combined 2: ')
    # print(row_data_combo_2)
    # print(f'combined 3: ')
    # print(row_data_combo_3)

    table1 = daff.PythonTableView(row_data_combo_1)  # daff needs a header in order to work correctly!
    table2 = daff.PythonTableView(row_data_combo_2)
    table3 = daff.PythonTableView(row_data_combo_3)

    # old version:
    # table1 = daff.PythonTableView([list(r.values()) for r in row_data_1])
    # table2 = daff.PythonTableView([list(r.values()) for r in row_data_2])

    # alignment = daff.Coopy.compareTables3(table3, table2, table1).align()  # 3 way: initial vs server vs saving

    # alignment = daff.Coopy.compareTables(table3,table2).align() #initial vs server
    alignment2 = daff.Coopy.compareTables(table2, table1).align()  # saving vs server
    # alignment2 = daff.Coopy.compareTables(table1, table2).align() #server vs saving
    # alignment = daff.Coopy.compareTables(table3,table1).align() #initial vs saving

    data_diff = []
    table_diff = daff.PythonTableView(data_diff)

    flags = daff.CompareFlags()

    # flags.allowDelete()
    # flags.allowUpdate()
    # flags.allowInsert()

    highlighter = daff.TableDiff(alignment2, flags)

    highlighter.hilite(table_diff)
    # hasDifference() should return true - and it does.
    if highlighter.hasDifference():
        current_app.logger.debug('HASDIFFERENCE')
        current_app.logger.debug(highlighter.getSummary().row_deletes)
    else:
        current_app.logger.debug('no difference found')
    diff2html = daff.DiffRender()
    diff2html.usePrettyArrows(False)
    diff2html.render(table_diff)
    table_diff_html = diff2html.html()

    # print(table_diff_html)
    # print(f'table 1 before patch test: ')
    # print(table1.toString())
    # patch test:
    # patcher = daff.HighlightPatch(table2,table_diff)
    # patcher.apply()
    # print(f'patch tester: ..................')
    # print(f'table1:')
    # print(table1.toString())
    # print(f'table2:')
    # print(table2.toString())
    # print(table2.toString())
    # print(f'table3:')
    # print(table3.toString())
    # table2String = table2.toString().strip() #no
    # todo: Task 1: turn MergeData into a Dict in order to post it to Github!
    # - use Janna's sheet builder example?
    # - post direct instead of going through Flask front-end?
    # table2String.strip()
    # table2Json = json.dumps(table2)
    # table2Dict = dict(todo: make this into a dict with id:0++ per row here!)
    # table2String = dict(table2String) #nope

    # merger test:
    # print(f'Merger test: ')
    merger = daff.Merger(table3, table2, table1, flags)  # (3initial, 1saving, 2server, flags)
    merger.apply()
    # print(f'table2:')
    # table2String = table2.toString()
    # print(table2String) #after merger

    data = table2.getData()  # merger table in list format
    # print(f'data: ')
    # print(json.dumps(data)) #it's a list.
    # convert to correct format (list of dicts):
    dataDict = []
    iter = -1
    for k in data:
        # add "id" value:
        iter = iter + 1
        dictT = {}
        if iter == 0:
            pass
            # print(f'header row - not using')
        else:
            dictT['id'] = iter  # add "id" with iteration
            for key, val in zip(row_header, k):
                # deal with conflicting val?

                dictT[key] = val
        # add to list:
        if iter > 0:  # not header - now empty dict
            dataDict.append(dictT)
            # print(f'update: ')
        # print(dataDict)

    # print(f'dataDict: ')
    # print(json.dumps(dataDict))
    # print(f'the type of dataDict is: ')
    # print(type(dataDict))

    # print(f'merger data:') #none
    # print(daff.DiffSummary().different) #nothing here?
    # mergerConflictInfo = merger.getConflictInfos()

    # print(f'Merger conflict infos: ')
    # print(f'table1:')
    # print(table1.toString())
    # print(f'table2:')
    # print(table2.toString())
    # print(f'table3:')
    # print(table3.toString())

    return table_diff_html, dataDict


@bp.route('/edit_external/<repo_key>/<path:folder_path>')
@verify_logged_in
def edit_external(repo_key, folder_path, github: GitHub, config: ConfigurationService):
    # print("edit_external reached")
    repository = config.get(repo_key)
    folder = folder_path
    spreadsheets = []
    directories = github.get(
        f'repos/{repository.full_name}/contents/{folder_path}'
    )
    for directory in directories:
        spreadsheets.append(directory['name'])
    # todo: need unique name for each? Or do we append to big array?
    # for spreadsheet in spreadsheets:
    #     print("spreadsheet: ", spreadsheet)

    sheet1, sheet2, sheet3 = spreadsheets
    (file_sha1, rows1, header1) = get_spreadsheet(github, repository.full_name, folder, sheet1)
    # not a spreadsheet but a csv file:
    (file_sha2, rows2, header2) = get_csv(github, repository.full_name, folder, sheet2)
    (file_sha3, rows3, header3) = get_csv(github, repository.full_name, folder, sheet3)
    return render_template('edit_external.html',
                           login=g.user.github_login,
                           repo_name=repo_key,
                           folder_path=folder_path,
                           spreadsheets=spreadsheets,  # todo: delete, just for test
                           rows1=json.dumps(rows1),
                           rows2=json.dumps(rows2),
                           rows3=json.dumps(rows3)
                           )
