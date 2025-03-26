from io import BytesIO
from typing import Union

import openpyxl
from flask_github import GitHub
from openpyxl.styles import PatternFill
from pyhornedowl import pyhornedowl

import ose.utils.github as github
from ose import constants
from ose.model.ExcelOntology import ExcelOntology
from ose.model.Term import UnresolvedTerm, Term
from ose.services.ConfigurationService import ConfigurationService
from ose.utils import get_spreadsheets, str_empty


def is_incomplete(term: Union[UnresolvedTerm, Term]):
    return str_empty(term.label) or len(term.sub_class_of) <= 0 or str_empty(term.definition())


def main(gh: GitHub, config: ConfigurationService, repo: str):
    repository = config.get(repo)
    full_repo = repository.full_name
    branch = repository.main_branch
    active_sheets = repository.indexed_files
    regex = "|".join(f"({r})" for r in active_sheets)

    externals_owl = "addicto_external.owl"
    externals_content = github.get_file(gh, full_repo, externals_owl).decode()

    external_ontology = ExcelOntology(externals_owl)
    ontology = pyhornedowl.open_ontology(externals_content, "rdf")
    for [p, d] in repository.prefixes:
        ontology.prefix_mapping.add_prefix(p, d)

    for c in ontology.get_classes():
        id = ontology.get_id_for_iri(c)
        labels = ontology.get_annotations(c, constants.RDFS_LABEL)

        if id is not None:
            for label in labels:
                external_ontology.add_term(Term(
                    id=id,
                    label=label,
                    origin=("<external>", -1),
                    relations=[],
                    sub_class_of=[],
                    equivalent_to=[],
                    disjoint_with=[]
                ))

    excel_files = get_spreadsheets(gh, full_repo, branch, include_pattern=regex)

    changed_files = 0

    for file in excel_files:
        has_changed = False
        content = github.get_file(gh, full_repo, file)
        o = ExcelOntology(file)
        o.import_other_excel_ontology(external_ontology)
        o.add_terms_from_excel(file, content)

        wb: openpyxl.workbook.Workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = wb.active

        header = [cell.value for cell in sheet[1] if cell.value]

        if "Label" not in header or "Curation status" not in header:
            continue

        c_label = header.index("Label")
        c_curation_status = header.index("Curation status")

        o.resolve()

        for row in sheet:
            label = row[c_label].value
            label = label.strip() if label is not None else None
            term = o._term_by_label(label)

            if term is not None and (str_empty(term.id) or term.id.startswith(repo.upper())):
                incomplete = is_incomplete(term)
                incomplete = incomplete or all(
                    p.is_unresolved() or
                    o._term_by_id(p.id) is not None and o._term_by_id(p.id).curation_status() == 'Pre-proposed'
                    for p in term.sub_class_of
                )
                if incomplete and row[c_curation_status].value != "Pre-proposed":
                    row[c_curation_status].value = "Pre-proposed"
                    for c in row:
                        c.fill = PatternFill(fgColor="ebfad0", fill_type="solid")

                    changed_files += 1
                    has_changed = True

        if has_changed:
            spreadsheet_stream = BytesIO()
            wb.save(spreadsheet_stream)
            msg = f"Updating {file.split('/')[-1]}\n\nSet terms to Pre-proposed"
            github.save_file(gh, full_repo, file, spreadsheet_stream.getvalue(), msg, branch)

    return f"Updated {changed_files} terms in {len(excel_files)} spreadsheets."
