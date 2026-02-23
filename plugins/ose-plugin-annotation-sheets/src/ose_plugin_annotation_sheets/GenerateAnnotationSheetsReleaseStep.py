import re
from typing import Dict, Optional

import networkx as nx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from flask_github import GitHub
from flask_sqlalchemy import SQLAlchemy

from ose.release.ReleaseStep import ReleaseStep
from ose.release.hierarchy import build_hierarchy
from ose.model.ReleaseScript import ReleaseScript
from ose.model.Result import Result
from ose.services.ConfigurationService import ConfigurationService

# Default color palette (cycled when no explicit color is provided)
_DEFAULT_PALETTE = [
    "4472C4",  # blue
    "ED7D31",  # orange
    "A5A5A5",  # grey
    "FFC000",  # gold
    "5B9BD5",  # light blue
    "70AD47",  # green
    "FF66CC",  # pink
    "7030A0",  # purple
    "0066FF",  # bright blue
    "00B050",  # emerald
]


def _color_shades(base_hex: str) -> tuple[str, str, str]:
    """Return three shades from a base hex color: darkest, medium, lightest.

    Shades are produced by blending towards white at 0%, 30%, and 55%.
    """
    r, g, b = int(base_hex[:2], 16), int(base_hex[2:4], 16), int(base_hex[4:6], 16)

    def blend(ratio: float) -> str:
        nr = int(r + (255 - r) * ratio)
        ng = int(g + (255 - g) * ratio)
        nb = int(b + (255 - b) * ratio)
        return f"{nr:02X}{ng:02X}{nb:02X}"

    return base_hex.upper(), blend(0.30), blend(0.55)


class GenerateAnnotationSheetsReleaseStep(ReleaseStep):
    def __init__(
        self,
        db: SQLAlchemy,
        gh: GitHub,
        release_script: ReleaseScript,
        release_id: int,
        tmp: str,
        config: ConfigurationService,
        *,
        included_files: list[str],
        intro_sheet: Optional[str] = None,
        sheets: Optional[Dict[str, dict]] = None,
        arms: Optional[list[str]] = None,
        annotation_columns: Optional[list[str]] = None,
        article_header: str = "Article 1 (citation)",
        output_filename: Optional[str] = None,
    ) -> None:
        super().__init__(db, gh, release_script, release_id, tmp, config)

        self._included_files = included_files
        self._intro_sheet = intro_sheet
        self._sheets_config = sheets or {}
        self._arms = arms or []
        self._annotation_columns = annotation_columns or []
        self._article_header = article_header
        self._output_filename = output_filename

    @classmethod
    def name(cls) -> str:
        return "ANNOTATION_SHEETS"

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _sheet_display_name(self, file_key: str) -> str:
        """Derive a human-readable sheet name from a file key."""
        cfg = self._sheets_config.get(file_key)
        if cfg and cfg.get("name"):
            return cfg["name"]
        # Strip repo prefix, title-case, replace underscores
        short = self._release_script.short_repository_name.lower()
        display = re.sub(rf"^{re.escape(short)}[_]?", "", file_key, flags=re.IGNORECASE)
        return display.replace("_", " ").strip().title()

    def _sheet_color(self, file_key: str, index: int) -> str:
        """Return the base hex color for a sheet."""
        cfg = self._sheets_config.get(file_key)
        if cfg and cfg.get("color"):
            color = cfg["color"].lstrip("#")
            return color
        return _DEFAULT_PALETTE[index % len(_DEFAULT_PALETTE)]

    def _total_annotation_cols(self) -> int:
        return len(self._arms) * len(self._annotation_columns)

    # ------------------------------------------------------------------
    # Workbook generation
    # ------------------------------------------------------------------

    def run(self) -> bool:
        result = Result(())

        files = [(k, f) for k, f in self._release_script.files.items() if k in self._included_files]
        self._total_items = len(files)

        if self._intro_sheet:
            # Download and use the intro spreadsheet as the base workbook (keeps its first sheet intact)
            self._download(self._intro_sheet)
            wb = openpyxl.load_workbook(self._local_name(self._intro_sheet))
            # Remove all sheets except the first
            for sheet_name in wb.sheetnames[1:]:
                wb.remove(wb[sheet_name])
        else:
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            if default_sheet is not None:
                wb.remove(default_sheet)

        for idx, (file_key, file) in enumerate(files):
            self._next_item(item=file.target.file, message="Generating annotation sheet for")

            G, ontology = build_hierarchy(file, self._artifacts(), self._local_name, self._repo_config.prefixes)
            sheet_name = self._sheet_display_name(file_key)
            # Excel sheet names are limited to 31 characters
            sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            base_color = self._sheet_color(file_key, idx)
            self._populate_sheet(ws, G, ontology, sheet_name, base_color)

        # Save
        short = self._release_script.short_repository_name
        if self._output_filename:
            file_name = self._output_filename
        else:
            file_name = f"{short}_data_extraction_template.xlsx"

        local_path = self._local_name(file_name)
        wb.save(local_path)
        self._store_artifact(local_path, file_name)

        self._set_release_result(result)
        return result.ok()

    def _populate_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        G: nx.DiGraph,
        ontology,
        sheet_name: str,
        base_color: str,
    ) -> None:
        height = nx.dag_longest_path_length(G) + 1 if len(G) > 0 else 1
        # Columns: ID | Level 1..N | Definition | [annotation cols per arm]
        hierarchy_cols = 1 + height  # ID + level labels
        definition_col = hierarchy_cols + 1
        total_hierarchy_and_def = definition_col
        annotation_start = total_hierarchy_and_def + 1

        n_annotation_cols = self._total_annotation_cols()
        total_cols = total_hierarchy_and_def + n_annotation_cols

        dark, medium, light = _color_shades(base_color)
        fill_dark = PatternFill(start_color=dark, end_color=dark, fill_type="solid")
        fill_medium = PatternFill(start_color=medium, end_color=medium, fill_type="solid")
        fill_light = PatternFill(start_color=light, end_color=light, fill_type="solid")
        font_bold = Font(bold=True)
        font_bold_white = Font(bold=True, color="FFFFFF")
        wrap_alignment = Alignment(wrap_text=True)

        # Determine if dark color needs white text
        r, g, b = int(dark[:2], 16), int(dark[2:4], 16), int(dark[4:6], 16)
        luminance = 0.299 * r + 0.587 * g + 0.114 * b
        header_font = font_bold_white if luminance < 160 else font_bold

        # ── Row 1: Article header over annotation columns ──
        # Hierarchy side is empty/merged, annotation side has the article header
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_dark
            cell.font = header_font
        if n_annotation_cols > 0:
            ws.cell(row=1, column=annotation_start).value = self._article_header
            if n_annotation_cols > 1:
                ws.merge_cells(
                    start_row=1, start_column=annotation_start,
                    end_row=1, end_column=total_cols,
                )
        if total_hierarchy_and_def > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_hierarchy_and_def)

        # ── Row 2: Ontology name over hierarchy + arm names over annotation columns ──
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = fill_medium
            cell.font = font_bold
        ws.cell(row=2, column=1).value = sheet_name
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_hierarchy_and_def)

        # Arm name headers
        cols_per_arm = len(self._annotation_columns)
        for arm_idx, arm_name in enumerate(self._arms):
            start_col = annotation_start + arm_idx * cols_per_arm
            ws.cell(row=2, column=start_col).value = arm_name
            if cols_per_arm > 1:
                ws.merge_cells(
                    start_row=2, start_column=start_col,
                    end_row=2, end_column=start_col + cols_per_arm - 1,
                )

        # ── Row 3: Column labels ──
        labels = ["ID"] + [f"Level {i + 1} label" for i in range(height)] + ["Definition"]
        for arm in self._arms:
            labels.extend(self._annotation_columns)

        for col_idx, label in enumerate(labels, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = label
            cell.fill = fill_light
            cell.font = font_bold
            cell.alignment = wrap_alignment

        # ── Row 4+: Data rows ──
        row_num = 4
        roots = [(n, d) for n, d in G.nodes(data=True) if G.in_degree(n) == 0]
        for root, root_data in roots:
            self._write_data_row(ws, row_num, ontology, root, root_data, 0, height)
            row_num += 1
            for c in nx.dfs_preorder_nodes(G, root):
                if c == root:
                    continue
                d = G.nodes[c]
                depth = nx.shortest_path_length(G, root, c)
                self._write_data_row(ws, row_num, ontology, c, d, depth, height)
                row_num += 1

        # ── Freeze panes: below headers, after ID column ──
        ws.freeze_panes = ws.cell(row=4, column=2).coordinate

        # ── Column widths ──
        ws.column_dimensions["A"].width = 14  # ID
        for i in range(height):
            col_letter = get_column_letter(2 + i)
            ws.column_dimensions[col_letter].width = 17  # Level labels
        ws.column_dimensions[get_column_letter(definition_col)].width = 50  # Definition
        for i in range(n_annotation_cols):
            col_letter = get_column_letter(annotation_start + i)
            ws.column_dimensions[col_letter].width = 13  # Annotation columns

    def _write_data_row(self, ws, row_num: int, ontology, node_iri: str, data: dict, depth: int, height: int) -> None:
        ws.cell(row=row_num, column=1).value = ontology.get_id_for_iri(node_iri)
        # Label at the appropriate depth column (column 2 + depth)
        label = data.get("label", node_iri)
        ws.cell(row=row_num, column=2 + depth).value = label
        # Definition column = 2 + height
        definition_col = 2 + height
        ws.cell(row=row_num, column=definition_col).value = data.get("definition", "")
