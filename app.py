# Copyright 2018 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License")
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import base64
# [START gae_python37_app]
import io
import json
import threading
import traceback
from datetime import date
from datetime import datetime

import daff
import openpyxl
from flask import Flask, request, g, session, redirect, url_for, render_template
from flask import jsonify
from flask_caching import Cache
from flask_cors import CORS  # enable cross origin request?
from flask_github import GitHub
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from OntologyDataStore import OntologyDataStore
from SpreadsheetSearcher import SpreadsheetSearcher
from config import *
from database.Base import db_session, init_db
from database.User import User
from guards.admin import verify_admin
from guards.verify_login import verify_logged_in
from utils.github import get_csv, get_spreadsheet

# setup sqlalchemy

logger = logging.getLogger(__name__)


# Create an app instance
class FlaskApp(Flask):
    def __init__(self, *args, **kwargs):
        super(FlaskApp, self).__init__(*args, **kwargs)
        self._activate_background_job()

    def _activate_background_job(self):
        init_db()


app = FlaskApp(__name__)
CORS(app)  # cross origin across all
app.config['CORS_HEADERS'] = 'Content-Type'
cors = CORS(app, resources={
    r"/api/*": {
        "origins": "*"
    }
})

app.config.from_object('config')
cache = Cache(app)
# set cache for each prefix in prefixes:    
for prefix in PREFIXES:
    cache.set("latestID" + prefix[0], 0)
# cache.set("latestID",0) #initialise caching
logger.info("cache initialised")

github = GitHub(app)
searcher = SpreadsheetSearcher(cache, app.config, github)
ontodb = OntologyDataStore(app.config)


@app.before_request
def before_request():
    g.user = None
    if 'user_id' in session:
        # print("user-id in session: ",session['user_id'])
        g.user = User.query.get(session['user_id'])


@app.after_request
def after_request(response):
    db_session.remove()
    # print("after_request is running")
    return response


@app.teardown_request
def teardown_request_func(error=None):
    try:
        db_session.remove()
    except Exception as e:
        logger.error(f"Error in teardown_request_func: {e}")
    if error:
        logger.error(str(error))


@github.access_token_getter
def token_getter():
    user = g.user
    if user is not None:
        return user.github_access_token


@app.route('/github-callback')
@github.authorized_handler
def authorized(access_token):
    next_url = request.args.get('next') or url_for('home')
    if access_token is None:
        logger.warning("Authorization failed.")
        return redirect(url_for('logout'))

    user = User.query.filter_by(github_access_token=access_token).first()
    if user is None:
        user = User(access_token)
    # Not necessary to get these details here
    # but it helps humans to identify users easily.
    g.user = user
    github_user = github.get('/user')
    user.github_id = github_user['id']
    user.github_login = github_user['login']
    user.github_access_token = access_token
    db_session.add(user)
    db_session.commit()

    session['user_id'] = user.id
    return redirect(next_url)


@app.route('/login')
def login():
    if session.get('user_id', None) is not None:
        session.pop('user_id', None)  # Could be stale
    return github.authorize(scope="user,repo")


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('loggedout'))


@app.route("/loggedout")
def loggedout():
    """
    Displays the page to be shown to logged out users.
    """
    return render_template("loggedout.html")


@app.route('/user')
@verify_logged_in
def user():
    return jsonify(github.get('/user'))


# Pages for the app
@app.route("/rebuild-index")
@verify_admin
def rebuild_index():
    sheets = searcher.rebuild_index()
    return ('<h1>Index rebuild</h1>'
            '<p>The index was rebuild successfully</p>'
            '<p><a href="/">Go back to main page</a></p>'
            '<h3> Files included in the index</h3>'
            '<ul>'
            f"{''.join(f'<li>{s}</li>' for s in sheets)}"
            '</ul>')

@app.route('/search', methods=['POST'])
@verify_logged_in
def search():
    searchTerm = request.form.get("inputText")
    repoName = request.form.get("repoName")
    searchResults = searchAcrossSheets(repoName, searchTerm)
    searchResultsTable = json.dumps(searchResults)
    return (json.dumps({"message": "Success",
                        "searchResults": searchResultsTable}), 200)


@app.route('/searchAssignedToMe', methods=['POST'])
@verify_logged_in
def searchAssignedToMe():
    initials = request.form.get("initials")
    logger.debug("Searching for initials: " + initials)
    repoName = request.form.get("repoName")
    # below is searching in "Label" column?
    searchResults = searchAssignedTo(repoName, initials)
    searchResultsTable = json.dumps(searchResults)
    return (json.dumps({"message": "Success",
                        "searchResults": searchResultsTable}), 200)


@app.route('/')
@app.route('/home')
@verify_logged_in
def home():
    repositories = app.config['REPOSITORIES']
    user_repos = repositories.keys()
    # Filter just the repositories that the user can see
    if g.user.github_login in USERS_METADATA:
        user_repos = USERS_METADATA[g.user.github_login]["repositories"]

    repositories = {k: v for k, v in repositories.items() if k in user_repos}

    return render_template('index.html',
                           login=g.user.github_login,
                           repos=repositories)


@app.route('/repo/<repo_key>')
@app.route('/repo/<repo_key>/<path:folder_path>')
@verify_logged_in
def repo(repo_key, folder_path=""):
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    directories = github.get(
        f'repos/{repo_detail}/contents/{folder_path}'
    )
    dirs = []
    spreadsheets = []
    # go to edit_external:
    if folder_path == 'imports':
        return redirect(url_for('edit_external', repo_key=repo_key, folder_path=folder_path))
    for directory in directories:
        if directory['type'] == 'dir':
            dirs.append(directory['name'])
        elif directory['type'] == 'file' and '.xlsx' in directory['name']:
            spreadsheets.append(directory['name'])
    if g.user.github_login in USERS_METADATA:
        user_initials = USERS_METADATA[g.user.github_login]["initials"]
    else:
        logger.info(f"The user {g.user.github_login} has no known metadata")
        user_initials = g.user.github_login[0:2]

    return render_template('repo.html',
                           login=g.user.github_login,
                           user_initials=user_initials,
                           repo_name=repo_key,
                           folder_path=folder_path,
                           directories=dirs,
                           spreadsheets=spreadsheets,
                           )


@app.route("/direct", methods=["POST"])
@verify_logged_in
def direct():
    if request.method == "POST":
        type = json.loads(request.form.get("type"))
        repo = json.loads(request.form.get("repo"))
        sheet = json.loads(request.form.get("sheet"))
        go_to_row = json.loads(request.form.get("go_to_row"))
    repoStr = repo['repo']
    sheetStr = sheet['sheet']
    url = '/edit' + '/' + repoStr + '/' + sheetStr
    session['type'] = type['type']
    session['label'] = go_to_row['go_to_row']
    session['url'] = url
    return ('success')


@app.route("/validate", methods=["POST"])
@verify_logged_in
def verify():
    if request.method == "POST":
        cell = json.loads(request.form.get("cell"))
        column = json.loads(request.form.get("column"))
        rowData = json.loads(request.form.get("rowData"))
        headers = json.loads(request.form.get("headers"))
        table = json.loads(request.form.get("table"))
        # check for blank cells under conditions first:
    blank = {}
    unique = {}
    returnData, uniqueData = checkBlankMulti(1, blank, unique, cell, column, headers, rowData, table)
    if len(returnData) > 0 or len(uniqueData) > 0:
        return (json.dumps({"message": "fail", "values": returnData, "unique": uniqueData}))
    return ('success')


@app.route("/generate", methods=["POST"])
@verify_logged_in
def generate():
    if request.method == "POST":
        repo_key = request.form.get("repo_key")
        rowData = json.loads(request.form.get("rowData"))
        values = {}
        ids = {}
        for row in rowData:
            nextIdStr = str(searcher.get_next_id(repo_key))
            fill_num = app.config['DIGIT_COUNT']
            if repo_key == "BCIO":
                fill_num = fill_num - 1
            else:
                fill_num = fill_num
            id = repo_key.upper() + ":" + nextIdStr.zfill(fill_num)
            ids["ID" + str(row['id'])] = str(row['id'])
            values["ID" + str(row['id'])] = id
        return (json.dumps({"message": "idlist", "IDs": ids, "values": values}))  # need to return an array
    return ('success')


# validation checks here: 

# recursive check each cell in rowData:
def checkBlankMulti(current, blank, unique, cell, column, headers, rowData, table):
    for index, (key, value) in enumerate(
            rowData.items()):  # todo: really, we need to loop here, surely there is a faster way?
        if index == current:
            if key == "Label" or key == "Definition" or key == "Parent" or key == "Sub-ontology" or key == "Curation status":
                if key == "Definition" or key == "Parent":
                    status = rowData.get("Curation status")  # check for "Curation status"
                    if (status):
                        if rowData["Curation status"] == "Proposed" or rowData["Curation status"] == "External":
                            pass
                        else:
                            if value.strip() == "":
                                blank.update({key: value})
                    else:
                        pass  # no "Curation status" column
                else:
                    if value.strip() == "":
                        blank.update({key: value})
                    else:
                        pass
            if key == "Label" or key == "ID" or key == "Definition":
                if checkNotUnique(value, key, headers, table):
                    unique.update({key: value})
    # go again:
    current = current + 1
    if current >= len(rowData):
        return (blank, unique)
    return checkBlankMulti(current, blank, unique, cell, column, headers, rowData, table)


def checkNotUnique(cell, column, headers, table):
    counter = 0
    cellStr = cell.strip()
    if cellStr == "":
        return False
    # if Label, ID or Definition column, check cell against all other cells in the same column and return true if same
    for r in range(len(table)):
        row = [v for v in table[r].values()]
        del row[0]  # remove extra numbered "id" column
        for c in range(len(headers)):
            if headers[c] == "ID" and column == "ID":
                if row[c].strip() == cellStr:
                    counter += 1
                    if counter > 1:  # more than one of the same
                        return True
            if headers[c] == "Label" and column == "Label":
                if row[c].strip() == cellStr:
                    counter += 1
                    if counter > 1:
                        return True
            if headers[c] == "Definition" and column == "Definition":
                if row[c].strip() == cellStr:
                    counter += 1
                    if counter > 1:
                        return True
    return False


@app.route('/edit/<repo_key>/<path:folder>/<spreadsheet>')
@verify_logged_in
def edit(repo_key, folder, spreadsheet):
    if session.get('label') == None:
        go_to_row = ""
    else:
        go_to_row = session.get('label')
        session.pop('label', None)

    if session.get('type') == None:
        type = ""
    else:
        type = session.get('type')
        session.pop('type', None)
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    (file_sha, rows, header) = get_spreadsheet(github, repo_detail, folder, spreadsheet)
    if g.user.github_login in USERS_METADATA:
        user_initials = USERS_METADATA[g.user.github_login]["initials"]
    else:
        logger.info(f"The user {g.user.github_login} has no known metadata")
        user_initials = g.user.github_login[0:2]
    # Build suggestions data:
    if repo_key not in ontodb.releases or date.today() > ontodb.releasedates[repo_key]:
        ontodb.parseRelease(repo_key)
    suggestions = ontodb.getReleaseLabels(repo_key)
    suggestions = list(dict.fromkeys(suggestions))

    return render_template('edit.html',
                           login=g.user.github_login,
                           user_initials=user_initials,
                           all_initials=ALL_USERS_INITIALS,
                           repo_name=repo_key,
                           folder=folder,
                           spreadsheet_name=spreadsheet,
                           header=json.dumps(header),
                           rows=json.dumps(rows),
                           file_sha=file_sha,
                           go_to_row=go_to_row,
                           type=type,
                           suggestions=json.dumps(suggestions)
                           )


@app.route('/download_spreadsheet', methods=['POST'])
@verify_logged_in
def download_spreadsheet():
    repo_key = request.form.get("repo_key")
    folder = request.form.get("folder")
    spreadsheet = request.form.get("spreadsheet")
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    url = github.get(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
    download_url = url['download_url']
    logger.debug(f"Downloading spreadsheet from {download_url}")
    return (json.dumps({"message": "Success",
                        "download_url": download_url}), 200)
    # return redirect(download_url) #why not?
    # r = requests.get(url)
    # strIO = StringIO.StringIO(r.content)
    # return send_file(strIO, as_attachment=True, attachment_filename={spreadsheet})

    # todo: get spreadsheet location and return it  f"repos/{repo_detail}/contents/{folder}/{spreadsheet}"
    # spreadsheet_file = github.get(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}");
    # spreadsheet_file = github.get(
    #     f'repos/{repo_detail}/contents/{folder}/{spreadsheet}'
    # )
    # base64_bytes = spreadsheet_file['content'].encode('utf-8')
    # decoded_data = base64.decodebytes(base64_bytes)
    # bytesIO = io.BytesIO(decoded_data)
    # wb = openpyxl.load_workbook(io.BytesIO(decoded_data))
    # sheet = wb.active
    # wb.save(spreadsheet)
    # bytesIO.seek(0)  # go to the beginning of the stream
    # #
    # return send_file(
    #     bytesIO,
    #     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #     attachment_filename=f"{spreadsheet}.xlsx",
    #     as_attachment=True,
    #     cache_timeout=0
    # )
    # return ( json.dumps({"message":"Success",
    #                             "spreadsheet_file": spreadsheet_file}), 200 )
    # return send_file(spreadsheet_file, as_attachment=True, attachment_filename=spreadsheet)
    # return redirect(f"https://raw.githubusercontent.com/{repo_key}/{folder}/{spreadsheet}?token={g.user.github_access_token}")


@app.route('/save', methods=['POST'])
@verify_logged_in
def save():
    repo_key = request.form.get("repo_key")
    folder = request.form.get("folder")
    spreadsheet = request.form.get("spreadsheet")
    row_data = request.form.get("rowData")
    initial_data = request.form.get("initialData")
    file_sha = request.form.get("file_sha").strip()
    commit_msg = request.form.get("commit_msg")
    commit_msg_extra = request.form.get("commit_msg_extra")
    overwrite = False
    overwriteVal = request.form.get("overwrite")
    if overwriteVal == "true":
        overwrite = True

    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    restart = False  # for refreshing the sheet (new ID's)
    try:
        initial_data_parsed = json.loads(initial_data)
        row_data_parsed = json.loads(row_data)
        # Get the data, skip the first 'id' column
        initial_first_row = initial_data_parsed[0]
        initial_header = [k for k in initial_first_row.keys()]
        del initial_header[0]
        # Sort based on label
        # What if 'Label' column not present?
        if 'Label' in initial_first_row:
            initial_data_parsed = sorted(initial_data_parsed, key=lambda k: k['Label'] if k['Label'] else "")
        else:
            logger.warning(
                "While saving: No Label column present, so not sorting this.")  # do we need to sort - yes, for diff!

        first_row = row_data_parsed[0]
        header = [k for k in first_row.keys()]
        del header[0]
        # Sort based on label
        # What if 'Label' column not present?
        if 'Label' in first_row:
            row_data_parsed = sorted(row_data_parsed, key=lambda k: k['Label'] if k['Label'] else "")
        else:
            logger.warning(
                "While saving: No Label column present, so not sorting this.")  # do we need to sort - yes, for diff!

        logger.debug(f"Got file_sha: {file_sha}")

        wb = openpyxl.Workbook()
        sheet = wb.active

        for c in range(len(header)):
            sheet.cell(row=1, column=c + 1).value = header[c]
            sheet.cell(row=1, column=c + 1).font = Font(size=12, bold=True)
        for r in range(len(row_data_parsed)):
            row = [v for v in row_data_parsed[r].values()]
            del row[0]  # Tabulator-added ID column
            for c in range(len(header)):
                sheet.cell(row=r + 2, column=c + 1).value = row[c]
                # Set row background colours according to 'Curation status'
                # These should be kept in sync with those used in edit screen
                # TODO add to config
                # What if "Curation status" not present?
                if 'Curation status' in first_row:
                    if row[header.index("Curation status")] == "Discussed":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="ffe4b5", fill_type="solid")
                    elif row[header.index("Curation status")] == "Ready":  # this is depreciated
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="98fb98", fill_type="solid")
                    elif row[header.index("Curation status")] == "Proposed":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="ffffff", fill_type="solid")
                    elif row[header.index("Curation status")] == "To Be Discussed":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="eee8aa", fill_type="solid")
                    elif row[header.index("Curation status")] == "In Discussion":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="fffacd", fill_type="solid")
                    elif row[header.index("Curation status")] == "Published":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="7fffd4", fill_type="solid")
                    elif row[header.index("Curation status")] == "Obsolete":
                        sheet.cell(row=r + 2, column=c + 1).fill = PatternFill(fgColor="2f4f4f", fill_type="solid")

            # Generate identifiers:
            if 'ID' in first_row:
                if not row[header.index("ID")]:  # blank
                    if 'Label' and 'Parent' and 'Definition' in first_row:  # make sure we have the right sheet
                        if row[header.index("Label")] and row[header.index("Parent")] and row[
                            header.index("Definition")]:  # not blank
                            # generate ID here:
                            nextIdStr = str(searcher.get_next_id(repo_key))
                            fill_num = app.config['DIGIT_COUNT']
                            if repo_key == "BCIO":
                                fill_num = fill_num - 1
                            else:
                                fill_num = fill_num
                            id = repo_key.upper() + ":" + nextIdStr.zfill(fill_num)
                            new_id = id
                            for c in range(len(header)):
                                if c == 0:
                                    restart = True
                                    sheet.cell(row=r + 2, column=c + 1).value = new_id

        # Create version for saving
        spreadsheet_stream = io.BytesIO()
        wb.save(spreadsheet_stream)

        # base64_bytes = base64.b64encode(sample_string_bytes)
        base64_bytes = base64.b64encode(spreadsheet_stream.getvalue())
        base64_string = base64_bytes.decode("ascii")

        # Create a new branch to commit the change to (in case of simultaneous updates)
        response = github.get(f"repos/{repo_detail}/git/ref/heads/master")
        if not response or "object" not in response or "sha" not in response["object"]:
            raise Exception(f"Unable to get SHA for HEAD of master in {repo_detail}")
        sha = response["object"]["sha"]
        branch = f"{g.user.github_login}_{datetime.utcnow().strftime('%Y-%m-%d_%H%M%S')}"
        logger.debug("About to try to create branch in %s", f"repos/{repo_detail}/git/refs")
        response = github.post(
            f"repos/{repo_detail}/git/refs", data={"ref": f"refs/heads/{branch}", "sha": sha},
        )
        if not response:
            raise Exception(f"Unable to create new branch {branch} in {repo_detail}")

        logger.debug("About to get latest version of the spreadsheet file %s",
                     f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        # Get the sha for the file
        (new_file_sha, new_rows, new_header) = get_spreadsheet(github, repo_detail, folder, spreadsheet)

        # Commit changes to branch (replace code with sheet)
        data = {
            "message": commit_msg,
            "content": base64_string,
            "branch": branch,
        }
        data["sha"] = new_file_sha
        logger.debug("About to commit file to branch %s", f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        response = github.put(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}", data=data)
        if not response:
            raise Exception(
                f"Unable to commit addition of {spreadsheet} to branch {branch} in {repo_detail}"
            )

        # Create a PR for the change
        logger.debug("About to create PR from branch", )
        response = github.post(
            f"repos/{repo_detail}/pulls",
            data={
                "title": commit_msg,
                "head": branch,
                "base": "master",
                "body": commit_msg_extra
            },
        )
        if not response:
            raise Exception(f"Unable to create PR for branch {branch} in {repo_detail}")
        pr_info = response['html_url']

        # Do not merge automatically if this file was stale as that will overwrite the other changes

        if new_file_sha != file_sha and not overwrite:
            logger.info("PR created and must be merged manually as repo file had changed")

            # Get the changes between the new file and this one:
            merge_diff, merged_table = getDiff(row_data_parsed, new_rows, new_header,
                                               initial_data_parsed)  # getDiff(saving version, latest server version, header for both)
            # update rows for comparison:
            (file_sha3, rows3, header3) = get_spreadsheet(github, repo_detail, folder, spreadsheet)
            # todo: delete transient branch here? Github delete code is a test for now.
            # Delete the branch again
            logger.debug("About to delete branch", f"repos/{repo_detail}/git/refs/heads/{branch}")
            response = github.delete(
                f"repos/{repo_detail}/git/refs/heads/{branch}")
            if not response:
                raise Exception(f"Unable to delete branch {branch} in {repo_detail}")
            return (
                json.dumps({
                    'Error': 'Your change was submitted to the repository but could not be automatically merged due to a conflict. You can view the change <a href="' \
                             + pr_info + '" target = "_blank" >here </a>. ', "file_sha_1": file_sha,
                    "file_sha_2": new_file_sha, "pr_branch": branch, "merge_diff": merge_diff,
                    "merged_table": json.dumps(merged_table), \
                    "rows3": rows3, "header3": header3}), 300  # 400 for missing REPO
            )
        else:
            # Merge the created PR
            logger.debug("About to merge created PR")
            response = github.post(
                f"repos/{repo_detail}/merges",
                data={
                    "head": branch,
                    "base": "master",
                    "commit_message": commit_msg
                },
            )
            if not response:
                raise Exception(f"Unable to merge PR from branch {branch} in {repo_detail}")

            # Delete the branch again
            logger.debug("About to delete branch %s", f"repos/{repo_detail}/git/refs/heads/{branch}")
            response = github.delete(
                f"repos/{repo_detail}/git/refs/heads/{branch}")
            if not response:
                raise Exception(f"Unable to delete branch {branch} in {repo_detail}")

        logger.info("Save succeeded.")
        # Update the search index for this file ASYNCHRONOUSLY (don't wait)
        thread = threading.Thread(target=searcher.update_index,
                                  args=(repo_key, folder, spreadsheet, header, row_data_parsed))
        thread.daemon = True  # Daemonize thread
        thread.start()  # Start the execution

        # Get the sha AGAIN for the file
        response = github.get(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        if not response or "sha" not in response:
            raise Exception(
                f"Unable to get the newly updated SHA value for {spreadsheet} in {repo_detail}/{folder}"
            )
        new_file_sha = response['sha']
        if restart:  # todo: does this need to be anywhere else also?
            return (json.dumps({"message": "Success",
                                "file_sha": new_file_sha}), 360)
        else:
            return (json.dumps({"message": "Success",
                                "file_sha": new_file_sha}), 200)

    except Exception as err:
        logger.error(err)
        traceback.print_exc()
        return (
            json.dumps({"message": "Failed",
                        "Error": format(err)}),
            400,
        )


@app.route('/keepalive', methods=['POST'])
@verify_logged_in
def keep_alive():
    # print("Keep alive requested from edit screen")
    return (json.dumps({"message": "Success"}), 200)


# todo: use this function to compare initial spreadsheet to server version - check for updates?
@app.route("/checkForUpdates", methods=["POST"])
def checkForUpdates():
    if request.method == "POST":
        repo_key = request.form.get("repo_key")
        folder = request.form.get("folder")
        spreadsheet = request.form.get("spreadsheet")
        # initialData = request.form.get("initialData") 
        old_sha = request.form.get("file_sha")
        repositories = app.config['REPOSITORIES']
        repo_detail = repositories[repo_key]
        spreadsheet_file = github.get(
            f'repos/{repo_detail}/contents/{folder}/{spreadsheet}'
        )
        file_sha = spreadsheet_file['sha']
        if old_sha == file_sha:
            return (json.dumps({"message": "Success"}), 200)
        else:
            return (json.dumps({"message": "Fail"}), 200)


@app.route('/openVisualiseAcrossSheets', methods=['POST'])
@verify_logged_in
def openVisualiseAcrossSheets():
    # build data we need for dotStr query (new one!)
    if request.method == "POST":
        idString = request.form.get("idList")
        repo = request.form.get("repo")
        idList = idString.split()
        ontodb.parseRelease(repo)
        # todo: do we need to support more than one repo at a time here?
        dotStr = ontodb.getDotForIDs(repo, idList).to_string()
        return render_template("visualise.html", sheet="selection", repo=repo, dotStr=dotStr)

    return ("Only POST allowed.")


# api:

@app.route('/api/openVisualiseAcrossSheets', methods=['POST'])
# @verify_logged_in # not enabled for /api/
def apiOpenVisualiseAcrossSheets():
    # build data we need for dotStr query (new one!)
    if request.method == "POST":
        idString = request.form.get("idList")
        repo = request.form.get("repo")
        idList = idString.split()
        ontodb.parseRelease(repo)
        dotStr = ontodb.getDotForIDs(repo, idList).to_string()
        # NOTE: APP_TITLE2 can't be blank - messes up the spacing
        APP_TITLE2 = "VISUALISATION"  # could model this on calling url here? Or something else..
        return render_template("visualise.html", sheet="selection", repo=repo, dotStr=dotStr, api=True,
                               APP_TITLE2=APP_TITLE2)


@app.route('/openVisualise', methods=['POST'])
@verify_logged_in
def openVisualise():
    curation_status_filters = ["", "External", "Proposed", "To Be Discussed", "In Discussion", "Discussed", "Published",
                               "Obsolete"]
    dotstr_list = []
    if request.method == "POST":
        repo = request.form.get("repo")
        sheet = request.form.get("sheet")
        table = json.loads(request.form.get("table"))
        indices = json.loads(request.form.get("indices"))
        try:
            filter = json.loads(request.form.get("filter"))
            if len(filter) > 0:
                pass
            else:
                filter = ""
        except Exception as err:
            filter = ""
            logger.error(str(err))
        if repo not in ontodb.releases:
            ontodb.parseRelease(repo)

        if len(indices) > 0:  # visualise selection
            # check if filter is greater than 1:
            if len(filter) > 1 and filter != "":  # multi-select:
                for i in range(0, 2):
                    ontodb.parseSheetData(repo, table)
                    dotStr = ontodb.getDotForSelection(repo, table, indices,
                                                       filter).to_string()  # filter is a list of strings here
            else:
                for filter in curation_status_filters:
                    # loop this twice to mitigate ID bug:
                    for i in range(0, 2):
                        ontodb.parseSheetData(repo, table)
                        dotStr = ontodb.getDotForSelection(repo, table, indices, filter).to_string()
                    # append dotStr to dotstr_list
                    dotstr_list.append(dotStr)  # all possible graphs
                # calculate default all values:
                filter = ""  # default
                for i in range(0, 2):
                    ontodb.parseSheetData(repo, table)
                    dotStr = ontodb.getDotForSelection(repo, table, indices, filter).to_string()

        else:
            # check if filter is greater than 1:
            if len(filter) > 1 and filter != "":  # multi-select:
                for i in range(0, 2):
                    ontodb.parseSheetData(repo, table)
                    dotStr = ontodb.getDotForSheetGraph(repo, table,
                                                        filter).to_string()  # filter is a list of strings here
                    # no dotstr_list here, just one dotStr
                    # dotstr_list.append(dotStr) #all possible graphs
            else:
                for filter in curation_status_filters:  # Visualise sheet
                    # loop this twice to mitigate ID bug:
                    for i in range(0, 2):
                        ontodb.parseSheetData(repo, table)
                        dotStr = ontodb.getDotForSheetGraph(repo, table, filter).to_string()
                    # append dotStr to dotstr_list
                    dotstr_list.append(dotStr)  # all possible graphs
                # calculate default all values:
                filter = ""  # default
                for i in range(0, 2):
                    ontodb.parseSheetData(repo, table)
                    dotStr = ontodb.getDotForSheetGraph(repo, table, filter).to_string()

        return render_template("visualise.html", sheet=sheet, repo=repo, dotStr=dotStr, dotstr_list=dotstr_list,
                               filter=filter)

    return ("Only POST allowed.")


# todo: below is never reached? 
@app.route('/visualise/<repo>/<sheet>')
@verify_logged_in  # todo: does this need to be disabled to allow cross origin requests? apparently not!
def visualise(repo, sheet):
    # print("reached visualise")
    return render_template("visualise.html", sheet=sheet, repo=repo)


@app.route('/openPat', methods=['POST'])
@verify_logged_in
def openPat():
    if request.method == "POST":
        repo = request.form.get("repo")
        # print("repo is ", repo)
        sheet = request.form.get("sheet")
        # print("sheet is ", sheet)
        table = json.loads(request.form.get("table"))
        # print("table is: ", table)
        indices = json.loads(request.form.get("indices"))
        # print("indices are: ", indices)

        if repo not in ontodb.releases:
            ontodb.parseRelease(repo)
        if len(indices) > 0:  # selection
            # ontodb.parseSheetData(repo,table)
            allIDS = ontodb.getIDsFromSelection(repo, table, indices)
            # print("got allIDS: ", allIDS)
        else:  # whole sheet..
            ontodb.parseSheetData(repo, table)
            allIDS = ontodb.getIDsFromSheet(repo, table)
            # todo: do we need to do above twice?

        # print("allIDS: ", allIDS)
        # remove duplicates from allIDS:
        allIDS = list(dict.fromkeys(allIDS))

        allData = ontodb.getMetaData(repo, allIDS)
        # print("allData: ", allData) 
        # print("dotStr is: ", dotStr)
        return render_template("pat.html", repo=repo, all_ids=allIDS, all_data=allData)  # todo: PAT.html

    return ("Only POST allowed.")


@app.route('/openPatAcrossSheets', methods=['POST'])
@verify_logged_in
def openPatAcrossSheets():
    # build data we need for dotStr query (new one!)
    if request.method == "POST":
        idString = request.form.get("idList")
        # print("idString is: ", idString)
        repo = request.form.get("repo")
        logger.debug("repo is %s", repo)
        idList = idString.split()
        logger.debug("idList: %s", idList)
        # for i in idList:
        #     print("i is: ", i)
        # indices = json.loads(request.form.get("indices"))
        # print("indices are: ", indices)
        ontodb.parseRelease(repo)
        # todo: do we need to support more than one repo at a time here?
        allIDS = ontodb.getRelatedIDs(repo, idList)
        # print("allIDS: ", allIDS)
        # remove duplicates from allIDS:
        allIDS = list(dict.fromkeys(allIDS))

        # todo: all experimental from here:
        allData = ontodb.getMetaData(repo, allIDS)
        # print("TEST allData: ", allData)

        # dotStr = ontodb.getDotForIDs(repo,idList).to_string()
        return render_template("pat.html", repo=repo, all_ids=allIDS, all_data=allData)  # todo: PAT.html

    return ("Only POST allowed.")


@app.route('/edit_external/<repo_key>/<path:folder_path>')
@verify_logged_in
def edit_external(repo_key, folder_path):
    # print("edit_external reached") 
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    folder = folder_path
    spreadsheets = []
    directories = github.get(
        f'repos/{repo_detail}/contents/{folder_path}'
    )
    for directory in directories:
        spreadsheets.append(directory['name'])
    # todo: need unique name for each? Or do we append to big array?
    # for spreadsheet in spreadsheets:
    #     print("spreadsheet: ", spreadsheet)

    sheet1, sheet2, sheet3 = spreadsheets
    (file_sha1, rows1, header1) = get_spreadsheet(github, repo_detail, folder, sheet1)
    # not a spreadsheet but a csv file:
    (file_sha2, rows2, header2) = get_csv(github, repo_detail, folder, sheet2)
    (file_sha3, rows3, header3) = get_csv(github, repo_detail, folder, sheet3)
    return render_template('edit_external.html',
                           login=g.user.github_login,
                           repo_name=repo_key,
                           folder_path=folder_path,
                           spreadsheets=spreadsheets,  # todo: delete, just for test
                           rows1=json.dumps(rows1),
                           rows2=json.dumps(rows2),
                           rows3=json.dumps(rows3)
                           )


@app.route('/save_new_ontology', methods=['POST'])
@verify_logged_in
def save_new_ontology():
    new_ontology = request.form.get("new_ontology")
    logger.debug("Received new Ontology: %s" + new_ontology)
    response = "test"
    return (json.dumps({"message": "Success",
                        "response": response}), 200)


@app.route('/update_ids', methods=['POST'])
@verify_logged_in
def update_ids():
    current_ontology = request.form.get("current_ontology")
    new_IDs = request.form.get("new_IDs")
    logger.debug("Received new IDs from : %s" + current_ontology)
    logger.debug("data is: %s", new_IDs)
    response = "test"
    return (json.dumps({"message": "Success",
                        "response": response}), 200)


# Internal methods


def getDiff(row_data_1, row_data_2, row_header, row_data_3):  # (1saving, 2server, header, 3initial)

    # print(f'the type of row_data_3 is: ')
    # print(type(row_data_3))        

    # sort out row_data_1 format to be the same as row_data_2
    new_row_data_1 = []
    for k in row_data_1:
        dictT = {}
        for key, val, item in zip(k, k.values(), k.items()):
            if (key != "id"):
                if (val == ""):
                    val = None
                # add to dictionary:
                dictT[key] = val
        # add to list:
        new_row_data_1.append(dictT)

        # sort out row_data_3 format to be the same as row_data_2
    new_row_data_3 = []
    for h in row_data_3:
        dictT3 = {}
        for key, val, item in zip(h, h.values(), h.items()):
            if (key != "id"):
                if (val == ""):
                    val = None
                # add to dictionary:
                dictT3[key] = val
        # add to list:
        new_row_data_3.append(dictT3)

    row_data_combo_1 = [row_header]
    row_data_combo_2 = [row_header]
    row_data_combo_3 = [row_header]

    row_data_combo_1.extend(
        [list(r.values()) for r in new_row_data_1])  # row_data_1 has extra "id" column for some reason???!!!
    row_data_combo_2.extend([list(s.values()) for s in row_data_2])
    row_data_combo_3.extend([list(t.values()) for t in new_row_data_3])

    # checking:
    # print(f'row_header: ')
    # print(row_header)
    # print(f'row_data_1: ')
    # print(row_data_1)
    # print(f'row_data_2: ')
    # print(row_data_2)
    # print(f'combined 1: ')
    # print(row_data_combo_1)
    # print(f'combined 2: ')
    # print(row_data_combo_2)
    # print(f'combined 3: ')
    # print(row_data_combo_3)

    table1 = daff.PythonTableView(row_data_combo_1)  # daff needs a header in order to work correctly!
    table2 = daff.PythonTableView(row_data_combo_2)
    table3 = daff.PythonTableView(row_data_combo_3)

    # old version:
    # table1 = daff.PythonTableView([list(r.values()) for r in row_data_1])
    # table2 = daff.PythonTableView([list(r.values()) for r in row_data_2])

    alignment = daff.Coopy.compareTables3(table3, table2, table1).align()  # 3 way: initial vs server vs saving

    # alignment = daff.Coopy.compareTables(table3,table2).align() #initial vs server
    alignment2 = daff.Coopy.compareTables(table2, table1).align()  # saving vs server
    # alignment2 = daff.Coopy.compareTables(table1, table2).align() #server vs saving
    # alignment = daff.Coopy.compareTables(table3,table1).align() #initial vs saving

    data_diff = []
    table_diff = daff.PythonTableView(data_diff)

    flags = daff.CompareFlags()

    # flags.allowDelete()
    # flags.allowUpdate()
    # flags.allowInsert()

    highlighter = daff.TableDiff(alignment2, flags)

    highlighter.hilite(table_diff)
    # hasDifference() should return true - and it does.
    if highlighter.hasDifference():
        logger.debug(f'HASDIFFERENCE')
        logger.debug(highlighter.getSummary().row_deletes)
    else:
        logger.debug(f'no difference found')
    diff2html = daff.DiffRender()
    diff2html.usePrettyArrows(False)
    diff2html.render(table_diff)
    table_diff_html = diff2html.html()

    # print(table_diff_html)
    # print(f'table 1 before patch test: ')
    # print(table1.toString()) 
    # patch test: 
    # patcher = daff.HighlightPatch(table2,table_diff)
    # patcher.apply()
    # print(f'patch tester: ..................')
    # print(f'table1:')
    # print(table1.toString())
    # print(f'table2:')
    # print(table2.toString())
    # print(table2.toString()) 
    # print(f'table3:')
    # print(table3.toString()) 
    # table2String = table2.toString().strip() #no
    # todo: Task 1: turn MergeData into a Dict in order to post it to Github!
    # - use Janna's sheet builder example? 
    # - post direct instead of going through Flask front-end? 
    # table2String.strip()
    # table2Json = json.dumps(table2)
    # table2Dict = dict(todo: make this into a dict with id:0++ per row here!)
    # table2String = dict(table2String) #nope

    # merger test: 
    # print(f'Merger test: ') 
    merger = daff.Merger(table3, table2, table1, flags)  # (3initial, 1saving, 2server, flags)
    merger.apply()
    # print(f'table2:')
    # table2String = table2.toString()
    # print(table2String) #after merger

    data = table2.getData()  # merger table in list format
    # print(f'data: ')
    # print(json.dumps(data)) #it's a list.
    # convert to correct format (list of dicts):
    dataDict = []
    iter = -1
    for k in data:
        # add "id" value:
        iter = iter + 1
        dictT = {}
        if iter == 0:
            pass
            # print(f'header row - not using')
        else:
            dictT['id'] = iter  # add "id" with iteration
            for key, val in zip(row_header, k):
                # deal with conflicting val?

                dictT[key] = val
        # add to list:
        if iter > 0:  # not header - now empty dict
            dataDict.append(dictT)
            # print(f'update: ')
        # print(dataDict)

    # print(f'dataDict: ')
    # print(json.dumps(dataDict))
    # print(f'the type of dataDict is: ')
    # print(type(dataDict))

    # print(f'merger data:') #none
    # print(daff.DiffSummary().different) #nothing here? 
    # mergerConflictInfo = merger.getConflictInfos()

    # print(f'Merger conflict infos: ')
    # print(f'table1:')
    # print(table1.toString())
    # print(f'table2:')
    # print(table2.toString()) 
    # print(f'table3:')
    # print(table3.toString()) 

    return (table_diff_html, dataDict)


def searchAcrossSheets(repo_name, search_string):
    searcherAllResults = searcher.search_for(repo_name, search_string=search_string)
    return searcherAllResults


def searchAssignedTo(repo_name, initials):
    searcherAllResults = searcher.search_for(repo_name, assigned_user=initials)
    return searcherAllResults


if __name__ == "__main__":  # on running python app.py

    app.run(debug=app.config["DEBUG"], port=8080)  # run the flask app

# [END gae_python37_app]
